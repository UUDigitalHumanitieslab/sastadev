import xlsxwriter
from openpyxl import load_workbook

from sastadev.config import SDLOGGER

eps = ''
defaultcolwidth = 15
cola, colb, colc, cold, cole, colf, colg, colh, coli, colj, colk, coll, colm, coln, colo, colp, colq, colr, cols, \
    colt, colu, colv, colw, colx, coly, colz = 0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, \
    21, 22, 23, 24, 25


def xlsx_writerow(sheet, rowctr, row, format=None, formats=[]):
    lrow = len(row)
    lformats = len(formats)
    for colctr in range(lrow):
        if format is None:
            if formats != [] and colctr < lformats and formats[colctr] is not None:
                sheet.write(rowctr, colctr, row[colctr], formats[colctr])
            else:
                sheet.write(rowctr, colctr, row[colctr])
        else:
            sheet.write(rowctr, colctr, row[colctr], format)


def writetable(tabel, ws, startrow=0, startcol=0, rhformat=None, chformat=None, cellformat=None):
    curcol = startcol
    currow = startrow
    for row in tabel:
        for el in row:
            if currow == startrow:
                theformat = rhformat
            elif curcol == startcol:
                theformat = chformat
            else:
                theformat = cellformat
            ws.write(currow, curcol, el, theformat)
            curcol += 1
        currow += 1
        curcol = startcol


def mkworkbook(outfullname, headers, allrows, sheetname='Sheet1', freeze_panes=None, formats=[]):
    workbook = xlsxwriter.Workbook(outfullname, {"strings_to_numbers": True})
    bold = workbook.add_format({'bold': True})

    realformats = []
    for fmt in formats:
        if fmt is None:
            realformats.append(fmt)
        else:
            realfmt = workbook.add_format(fmt)
            realformats.append(realfmt)

    worksheet1 = workbook.add_worksheet(sheetname)

    # worksheet1
    if freeze_panes is not None:
        (r, c) = freeze_panes
        worksheet1.freeze_panes(r, c)

    colctr = 0
    if headers != []:
        for val in headers[-1]:
            if val is None:
                cval = ''
            else:
                cval = str(val)
            colwidth = len(cval) if len(cval) > defaultcolwidth else defaultcolwidth
            worksheet1.set_column(colctr, colctr, colwidth)
            colctr += 1

    rowctr = 0
    for header in headers:
        xlsx_writerow(worksheet1, rowctr, header, format=bold)
        rowctr += 1

    for row in allrows:
        xlsx_writerow(worksheet1, rowctr, row, formats=realformats)
        rowctr += 1

    worksheet1.autofilter(0, 0, rowctr, colctr)
    return workbook


def adaptformats(formats, workbook):
    realformats = []
    for fmt in formats:
        if fmt is None:
            realformats.append(fmt)
        else:
            realfmt = workbook.add_format(fmt)
            realformats.append(realfmt)
    return realformats


def add_worksheet(workbook, headers, allrows, sheetname='Sheet2', freeze_panes=None, formats=[]):
    bold = workbook.add_format({'bold': True})

    realformats = adaptformats(formats, workbook)

    worksheet1 = workbook.add_worksheet(sheetname)

    # worksheet1
    if freeze_panes is not None:
        (r, c) = freeze_panes
        worksheet1.freeze_panes(r, c)

    if headers != []:
        colctr = 0
        for val in headers[-1]:
            if val is None:
                cval = ''
            else:
                cval = str(val)
            colwidth = len(cval) if len(cval) > defaultcolwidth else defaultcolwidth
            worksheet1.set_column(colctr, colctr, colwidth)
            colctr += 1

    rowctr = 0
    for header in headers:
        xlsx_writerow(worksheet1, rowctr, header, format=bold)
        rowctr += 1

    for row in allrows:
        xlsx_writerow(worksheet1, rowctr, row, formats=realformats)
        rowctr += 1

    worksheet1.autofilter(0, 0, rowctr, colctr)
    return worksheet1


def getxlsxdata(fullname, headerrow=0, sheetname=None):
    data = []
    header = []
    try:
        wb = load_workbook(fullname)
    except:
        return (header, data)
    wsns = wb.sheetnames
    if wsns != []:
        if sheetname is None:
            ws = wb[wsns[0]]
        else:
            ws = wb[sheetname]
    else:
        SDLOGGER.error("Error reading {}. No worksheets found\n".format(fullname))
    rowctr = -1
    for row in ws.iter_rows():
        rowctr += 1
        #preheader
        if rowctr < headerrow:
            continue
        # header
        elif rowctr == headerrow:
            header = [eps if cell.value is None else cell.value for cell in row]
        else:
            valuerow = [eps if cell.value is None else cell.value for cell in row]
            data.append(valuerow)
    wb.close()
    return header, data


def write2excel(datadict, header, filename):
    data = [list(key) + datadict[key] for key in datadict]
    workbook = mkworkbook(filename, [header], data)
    workbook.close()